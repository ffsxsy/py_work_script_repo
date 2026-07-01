#!/usr/bin/env python3
"""从 McuCanMap.xlsx「TX CAN-A」生成测量响应 C 数组（格式同 meas_example.c）。

CAN ID 范围与解析规则见 ``docs/GEN_DSP_MEAS_RESP.md``（当前 0x1A80–0x1AA1）。

用法:
    python gen_dsp_meas_resp_from_xlsx.py
    python gen_dsp_meas_resp_from_xlsx.py --check-only
"""

from __future__ import annotations

import argparse
import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "McuCanMap.xlsx"
OUT_PATH = SCRIPT_DIR / "c_struct_meas.c"
SHEET = "TX CAN-A"
ID_RE = re.compile(r"0x([0-9A-Fa-f]+)ssdd", re.IGNORECASE)
CAN_ID_MIN = 0x1A80
CAN_ID_MAX = 0x1AA1
# 输出为 C 注释块（与 meas_example.c 一致，仍保留在表中便于对照 xlsx）
COMMENTED_CAN_IDS: Final[frozenset[int]] = frozenset({0x1A92})


@dataclass(slots=True)
class FieldValue:
    enum: str
    label: str
    ctype: str
    offset: int
    length: int
    scale: float
    source: str  # matched | fallback


@dataclass(slots=True)
class FramePacket:
    can_id: int
    names: list[str]
    details: list[list[object]]


def parse_can_id(cell: object) -> int | None:
    if not cell or not isinstance(cell, str):
        return None
    m = ID_RE.match(cell.strip())
    if not m:
        return None
    return int(m.group(1), 16)


def is_frame_layout(row: list[object]) -> bool:
    if len(row) < 10:
        return False
    payload = row[2:10]
    for i in range(0, 8, 2):
        if not isinstance(payload[i], str):
            return False
        if payload[i + 1] is not None:
            return False
    return True


def norm_label(name: str) -> str:
    base = name.strip()
    if "(" in base:
        base = base.split("(", 1)[0].strip()
    return base


def canonical_label(name: str) -> str:
    """用于跨表头/明细匹配字段名：仅保留字母数字并转小写。"""
    base = norm_label(name).lower()
    base = base.replace("p.u.", "").replace("p.u", "")
    return "".join(ch for ch in base if ch.isalnum())


def is_measure_header_row(row: list[object]) -> bool:
    return (
        isinstance(row[0], str)
        and row[0].strip().lower() == "node"
        and isinstance(row[1], str)
        and row[1].strip().lower() == "name"
        and isinstance(row[2], str)
        and row[2].strip().lower() == "variable"
        and isinstance(row[3], str)
        and row[3].strip().lower() == "default"
        and isinstance(row[4], str)
        and row[4].strip().lower() == "factor"
        and isinstance(row[5], str)
        and row[5].strip().lower() == "min"
        and isinstance(row[6], str)
        and row[6].strip().lower() == "max"
    )


def is_debug_trace_marker_row(row: list[object]) -> bool:
    return isinstance(row[0], str) and "Debug trace variable default, factor and range" in row[0]


def _object_to_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def parse_scale(detail: list[object]) -> float | None:
    return _object_to_float(detail[4])


def infer_c_type_from_min(detail: list[object], fallback: str) -> str:
    """最小值为负 -> int16_t；最小值非负 -> uint16_t。"""
    min_v = _object_to_float(detail[5])
    if min_v is None:
        return fallback
    return "int16_t" if min_v < 0 else "uint16_t"


def format_scale(f: float) -> str:
    if f.is_integer():
        return f"{int(f)}.0F"
    return f"{f}F"


def collect_frames(ws: Any) -> list[tuple[int, list[str]]]:
    frames: list[tuple[int, list[str]]] = []
    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        r = list(row)
        nid = parse_can_id(r[0])
        if nid is None or not (CAN_ID_MIN <= nid <= CAN_ID_MAX):
            continue
        if not is_frame_layout(r):
            continue
        names = [str(r[2 + j]) for j in range(0, 8, 2)]
        frames.append((nid, names))
    return frames


def collect_detail_groups_by_id(ws: Any) -> dict[int, list[list[list[object]]]]:
    """每组 = 一条显式 Node 明细行 + 后续空 Node 继承行。"""
    groups_by_id: dict[int, list[list[list[object]]]] = {}
    current_detail_id: int | None = None
    current_group: list[list[object]] | None = None
    in_measure_section = False

    def flush_group() -> None:
        nonlocal current_group, current_detail_id
        if current_group is None or current_detail_id is None:
            return
        groups_by_id.setdefault(current_detail_id, []).append(current_group)
        current_group = None

    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        r = list(row)
        if not in_measure_section:
            if not is_measure_header_row(r):
                continue
            in_measure_section = True
            continue
        if is_debug_trace_marker_row(r):
            flush_group()
            break

        nid = parse_can_id(r[0])
        if nid is not None:
            flush_group()
            if is_frame_layout(r):
                continue
            if CAN_ID_MIN <= nid <= CAN_ID_MAX:
                current_detail_id = nid
                current_group = [r]
            else:
                current_detail_id = None
                current_group = None
            continue

        if current_detail_id is None or current_group is None:
            continue
        if isinstance(r[1], str):
            current_group.append(r)

    flush_group()
    return groups_by_id


def attach_details_to_frames(
    frames: list[tuple[int, list[str]]],
    detail_groups_by_id: dict[int, list[list[list[object]]]],
) -> list[FramePacket]:
    packets: list[FramePacket] = []
    group_idx: dict[int, int] = {}
    for nid, names in frames:
        groups = detail_groups_by_id.get(nid, [])
        gi = group_idx.get(nid, 0)
        if gi < len(groups):
            details = groups[gi]
            group_idx[nid] = gi + 1
        else:
            details = []
        packets.append(FramePacket(can_id=nid, names=names, details=details))
    return packets


def _detail_name(detail_row: list[object]) -> str | None:
    if len(detail_row) < 2:
        return None
    dname = detail_row[1]
    if not isinstance(dname, str):
        return None
    s = dname.strip()
    return s if s else None


def resolve_fields(packet: FramePacket) -> list[FieldValue]:
    default_scale = 0.125
    default_type = "int16_t"
    if packet.details:
        ps = parse_scale(packet.details[0])
        if ps is not None:
            default_scale = ps

    detail_map: dict[str, list[object]] = {}
    for d in packet.details:
        b1 = d[1]
        if isinstance(b1, str):
            detail_map[canonical_label(b1)] = d

    fields: list[FieldValue] = []
    for idx, label in enumerate(packet.names):
        off = idx * 2
        enum = f"kDSP_{packet.can_id:04X}_param{idx + 1}"
        ctype = default_type
        scale = default_scale
        source = "fallback"
        detail_row: list[object] | None = None

        if idx < len(packet.details):
            detail_row = packet.details[idx]
            dname = _detail_name(detail_row)
            if dname is not None:
                label = dname
        else:
            detail_row = detail_map.get(canonical_label(label))

        if detail_row is not None:
            parsed_scale = parse_scale(detail_row)
            scale = default_scale if parsed_scale is None else parsed_scale
            ctype = infer_c_type_from_min(detail_row, default_type)
            source = "matched"

        fields.append(
            FieldValue(
                enum=enum,
                label=label,
                ctype=ctype,
                offset=off,
                length=2,
                scale=scale,
                source=source,
            )
        )
    return fields


def _as_c_comment(lines: list[str]) -> list[str]:
    commented: list[str] = []
    for line in lines:
        if not line.strip():
            commented.append(line)
        elif line.startswith("    "):
            commented.append("    // " + line[4:])
        else:
            commented.append("// " + line)
    return commented


def emit_message(packet: FramePacket) -> list[str]:
    fields = resolve_fields(packet)

    enum_w = max(len(f.enum) for f in fields)
    name_w = max(len(f'"{f.label}"') for f in fields)
    type_w = max(len(f'"{f.ctype}"') for f in fields)
    off_w = max(len(str(f.offset)) for f in fields)
    len_w = max(len(str(f.length)) for f in fields)
    scale_w = max(len(format_scale(f.scale)) for f in fields)

    bodies = [
        f"{f.enum.ljust(enum_w)}, "
        f"{f'"{f.label}"'.ljust(name_w)}, "
        f"{f'"{f.ctype}"'.ljust(type_w)}, "
        f"{str(f.offset).rjust(off_w)}, "
        f"{str(f.length).rjust(len_w)}, "
        f"{format_scale(f.scale).rjust(scale_w)}"
        for f in fields
    ]

    block: list[str] = []
    block.append(f"    {{0x{packet.can_id:04X},")
    block.append(f"     {len(fields)},")
    if len(bodies) == 1:
        block.append("     {{" + bodies[0] + "}" + "}},")
    else:
        block.append("     {{" + bodies[0] + "},")
        for b in bodies[1:-1]:
            block.append(f"      {{{b}}},")
        block.append("      {" + bodies[-1] + "}" + "}},")
    return block


def summarize_consistency(packets: list[FramePacket]) -> tuple[int, int, list[str]]:
    matched = 0
    fallback = 0
    fallback_lines: list[str] = []

    for p in packets:
        for f in resolve_fields(p):
            if f.source == "matched":
                matched += 1
            else:
                fallback += 1
                fallback_lines.append(
                    f"0x{p.can_id:04X} | {f.enum} | {f.label} | factor={format_scale(f.scale)}"
                )
    return matched, fallback, fallback_lines


def generate_lines(packets: list[FramePacket]) -> list[str]:
    lines: list[str] = []
    lines.append(
        "/*\n"
        " * 由 McuCanMap.xlsx「TX CAN-A」自动生成（CAN ID 0x1A80–0x1AA1，Node 形如 0x####ssdd）。\n"
        " * 数组元素按报文 CAN ID 升序排列；同 ID 重复项保持表中顺序。\n"
        " * 枚举占位符为 kDSP_<CANID>_param1 … paramN，请在头文件中定义或映射为实际枚举。\n"
        " * 参数名 / factor / 类型优先取自测量明细区（与帧布局 byte 列不一致时以明细为准）；\n"
        " * 类型由该行 min 是否小于 0 判定（负 -> int16_t，非负 -> uint16_t）。\n"
        " * 0x1A92 整帧输出为注释块（与 meas_example.c 一致）。\n"
        " */"
    )
    lines.append("meas_resp_value_t dsp_meas_resp_content_array[] = {")
    for packet in packets:
        block = emit_message(packet)
        if packet.can_id in COMMENTED_CAN_IDS:
            block = _as_c_comment(block)
        lines.extend(block)
    lines.append("};")
    lines.append("")
    return lines


def build_packets(xlsx_path: Path | None = None) -> list[FramePacket]:
    path = xlsx_path or XLSX_PATH
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*Workbook contains no default style.*",
            category=UserWarning,
        )
        wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[SHEET]
        frames = collect_frames(ws)
        detail_groups = collect_detail_groups_by_id(ws)
    finally:
        wb.close()

    valid_ids = set(detail_groups.keys())
    frames = [f for f in frames if f[0] in valid_ids]
    frames.sort(key=lambda x: x[0])
    return attach_details_to_frames(frames, detail_groups)


def write_meas_struct(
    out_path: Path | None = None,
    xlsx_path: Path | None = None,
) -> tuple[Path, int]:
    """生成 c_struct_meas.c，返回 (输出路径, 报文条数)。"""
    packets = build_packets(xlsx_path)
    target = out_path or OUT_PATH
    target.write_text("\n".join(generate_lines(packets)), encoding="utf-8")
    return target, len(packets)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从 McuCanMap.xlsx TX CAN-A 生成 dsp_meas_resp_content_array"
    )
    parser.add_argument("--check-only", action="store_true", help="只输出一致性统计，不写文件")
    parser.add_argument("--fallback-limit", type=int, default=20, help="回退项展示条数")
    parser.add_argument("-o", "--output", type=Path, default=OUT_PATH, help="输出 .c 路径")
    args = parser.parse_args()

    packets = build_packets()
    matched, fallback, fallback_lines = summarize_consistency(packets)
    total = matched + fallback

    print(f"字段匹配: total={total} matched={matched} fallback={fallback}")
    if fallback_lines:
        print("回退项样例:")
        for line in fallback_lines[: args.fallback_limit]:
            print("  " + line)

    if args.check_only:
        return

    out, count = write_meas_struct(args.output)
    print(f"已生成 {out}，共 {count} 条 CAN 报文")


if __name__ == "__main__":
    main()
