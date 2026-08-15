"""Verify Modbus discrete inputs & input registers against rbms config presets."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from pymodbus.client import ModbusTcpClient

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT / "src"))

from rbms_tcp_sim.codec import physical_to_raw  # noqa: E402
from rbms_tcp_sim.matrix_config.csv_common import load_matrix_csv  # noqa: E402
from rbms_tcp_sim.matrix_config.profiles import get_profile  # noqa: E402

DEFAULT_HOST = "192.168.1.136"
DEFAULT_PORT = 1502
DEFAULT_SLAVE = 1
INPUT_RACK_STRIDE = 3000
DISCRETE_RACK_STRIDE = 200
MAX_REGS_PER_READ = 120
XLSX = _ROOT / "docs/BMS2.0_BMS_EMS_IP_TCP_Register_Table_V1.0.13.xlsx"

# 点表 offset 351–377 L 列为空；Mapping 参考：
#   - SystemConfiguration_BMS20_RBMS.xlsm → FaultList（RBMS 位图，EMS_DISC_RBMS_BIT）
#   - 网关/固件离散输入映射表（见 docs；EMS_DISC_M_FAULT = BBMS M 核，非 rbms_fault.csv）
#
# 等级用语（点表「轻/中/重度」↔ FaultLevel）：轻度=3级  中度=2级  重度=1级
# Modbus 簇1 地址 = 200 + offset。
RACK_DISCRETE_SUPPLEMENT: dict[int, str] = {
    351: "ACIS_AEROSOL_FAULT",  # RBMS_BIT #163
    352: "ACIS_ESTOP_FAULT",  # RBMS_BIT #160
    353: "ACIS_AEROSOL2_FAULT",  # RBMS_BIT #164
    354: "kbmsMFault_FireModerateAlarm",  # M_FAULT；RBMS 无 FID；占位名取自固件
    355: "ACIS_FIRE_EXTINGUISHER_ALERT_FAULT",  # RBMS_BIT #142；固件 FireL1
    356: "ACIS_QF1_BREAKER_FAULT",  # RBMS_BIT #162
    357: "ACIS_WATER_FLOOD_FAULT",  # RBMS_BIT #161
    358: "kRbmsFault_EMSCommLost",  # RBMS_BIT；固件有、RBMS xlsm 无 FID（见歧义说明）
    359: "kbmsMFault_EnergyMeterCommLost",  # M_FAULT
    360: "EMCR_PCS_TO_BMS_COMMUNICATION_LOST_FAULT",  # RBMS_BIT #167
    361: "kbmsMFault_UPSCommLost",  # M_FAULT
    362: "kbmsMFault_WaterCoolLvl3Fault",  # M_FAULT；点表轻度=3级
    363: "kbmsMFault_WaterCoolLvl2Fault",  # M_FAULT；点表中度=2级
    364: "kbmsMFault_WaterCoolLvl1Fault",  # M_FAULT；点表重度=1级
    365: "kbmsMFault_WaterCoolModeMismatch",  # M_FAULT
    366: "kbmsMFault_TMSCommLost",  # M_FAULT
    367: "EMCR_IOModule_TO_BMS_COMMUNICATION_LOST_FAULT",  # RBMS_BIT #166
    368: "EMCR_DEHUMIDIFIER_TO_BMS_COMMUNICATION_LOST_FAULT",  # RBMS_BIT #165
    369: "ACIS_COOL_FAN_FAULT",  # RBMS_BIT #157
    370: "ACIS_SMOKE_FEEDBACK_FAULT",  # RBMS_BIT #158
    371: "ACIS_TEMP_FEEDBACK_FAULT",  # RBMS_BIT #159
    372: "BTCS_CANMSG_TIMEOUT_FAULT",  # RBMS_BIT #14
    373: "BTCS_CANBATI_IMPLY_FAULT",  # RBMS_BIT #15
    374: "BTCS_CANBATI_ZERODRIFT_FAULT",  # RBMS_BIT #16
    375: "BTCS_CANMSG_CRCRES_FAULT",  # RBMS_BIT #17
    376: "BTCS_HALLSENS_OVERRANGE_FAULT",  # RBMS_BIT #169；点表写「差异过大」见歧义
    377: "HVSH_MAIN_POSITIVE_BUS_BROKEN_FAULT",  # RBMS_BIT #156
}

# 仅 _lookup_fid 可识别的 RBMS FaultName（不含 M_FAULT / 固件占位名）
RACK_DISCRETE_SUPPLEMENT_RBMS_ONLY: dict[int, str] = {
    k: v
    for k, v in RACK_DISCRETE_SUPPLEMENT.items()
    if not v.startswith("kbmsMFault_") and not v.startswith("kRbmsFault_")
}

# Excel SOX 段 Mapping 列为空，按 offset 对齐 config 信号名
SOX_BY_OFFSET: dict[int, tuple[str, str] | None] = {
    2214: ("soxdebug1", "ScCSPC_FullChEnaNbr"),
    2215: ("soxdebug1", "ScCSPC_FullDischEnaNbr"),
    2216: ("soxdebug1", "ScHIST_HistAccuChCapAh"),
    2217: ("soxdebug1", "ScHIST_HistAccuChCapAh"),
    2218: ("soxdebug1", "ScHIST_HistAccuDischCapAh"),
    2219: ("soxdebug1", "ScHIST_HistAccuDischCapAh"),
    2220: ("soxdebug1", "ScSGPC_LstSleepDurTiMin"),
    2221: ("soxdebug1", "ScSGPC_LstSleepDurTiMin"),
    2222: ("soxdebug1", "ScSGPC_HisSleepTiMin"),
    2223: ("soxdebug1", "ScSGPC_HisSleepTiMin"),
    2224: ("soxdebug1", "ScSGPC_HisRunTiMin"),
    2225: ("soxdebug1", "ScSGPC_HisRunTiMin"),
    2226: ("soxdebug1", "ScHIST_CycNbr"),
    2227: ("soxdebug1", "BSWSAllSOHCalIndicator"),
    2228: ("soxdebug1", "BSWSAllSOHCalValPct"),
    2229: ("soxdebug1", "ScBTCS_RackChrgCapmAh"),
    2230: ("soxdebug1", "ScBTCS_RackChrgCapmAh"),
    2231: ("soxdebug1", "ScBTCS_RackDsChrgCapmAh"),
    2232: ("soxdebug1", "ScBTCS_RackDsChrgCapmAh"),
    2233: ("soxdebug1", "BSWSRSOXHistInfoIndicator"),
    2234: ("soxdebug1", "BSWSAllSOCCalIndicator"),
    2235: ("soxdebug1", "BSWSAllSOCCalValPct"),
    2236: ("soxdebug1", "ScSGPC_MaxCellVMdulTDegC"),
    2237: ("soxdebug1", "ScSGPC_MinCellVMdulTDegC"),
    2238: None,
    2239: ("soxdebug2", "ScSOHA_RealSysSOHCPct"),
    2240: ("soxdebug2", "ScSOHA_DFCLCapResultAh"),
    2241: ("soxdebug2", "ScSOHA_DFCLPointSOCPct"),
    2242: ("soxdebug2", "ScSOHA_DFCLPointStats"),
    2243: ("soxdebug2", "ScSOHA_DFCLPointCapAh"),
    2244: ("soxdebug2", "ScSOHA_DFCLPointTimeDay"),
    2245: ("soxdebug2", "ScSOHA_MFCLCapResultAh"),
    2246: ("soxdebug2", "ScSOHA_MFCLTargtValueVmax"),
    2247: ("soxdebug2", "ScSOHA_MFCLTargtValueVmin"),
    2248: ("soxdebug2", "SaSOCA_MaxMinSOCState1"),
    2249: ("soxdebug2", "SaSOCA_MaxMinSOCState2"),
    2250: ("soxdebug2", "ScSOHA_RealSysCapAh_dbug"),
    2251: ("soxdebug2", "ScSOCA_RealSysSOCPct_dbug"),
    2252: ("soxdebug2", "ScSOCA_DispSOCState"),
    2253: None,
}

INPUT_SEGMENTS: tuple[tuple[int, int, str], ...] = (
    (100, 45, "SumInfo 概要 100-144"),
    (201, 100, "电芯电压 201-300"),
    (901, 100, "模组温度 901-1000"),
    (1601, 100, "极柱温度 1601-1700"),
    (1761, 16, "Pack连接器温度 1761-1776"),
    (2201, 53, "SumInfo/SOX 2201-2253"),
)

PROFILE_BY_SRC = {
    "RBMS_SumInfo_Data": "suminfo",
    "RBMS_Volt_Data": "volt",
    "RBMS_Temp_Data": "temp",
    "RBMS_SOXdebugData1": "soxdebug1",
    "RBMS_SOXdebugData2": "soxdebug2",
}

_RESERVED_RE = re.compile(r"预留|\breserved\d*\b", re.IGNORECASE)


def _is_reserved_row(field: object, map_sig: object = None) -> bool:
    """点表字段名 / Mapping 含「预留」「Reserved」时不参与对比。"""
    if _RESERVED_RE.search(str(field or "")):
        return True
    return bool(map_sig and _RESERVED_RE.search(str(map_sig)))


def _is_reserved_addr_expr(expr: object) -> bool:
    """点表地址为区间预留（如 145 to 200、58-69）时不参与对比。"""
    s = str(expr or "").lower()
    return " to " in s or re.search(r"\d+\s*-\s*\d+", s) is not None


@dataclass(frozen=True)
class RawSignal:
    raw: int
    is_u32: bool


def _load_fault_bits() -> dict[int, int]:
    p = get_profile("fault")
    settings = load_matrix_csv(p.default_csv, skip_signals=p.skip_signals)
    return {
        int(s.signal.split("_")[-1]) - 1: int(s.value)
        for s in settings.signals
        if s.signal.startswith("RBMS_Fault_")
    }


def _load_name_to_fid() -> dict[str, int]:
    fault_md = (_ROOT / "docs/文档二次整理/SystemConfiguration_BMS20_RBMS-FaultList.md").read_text(
        encoding="utf-8"
    )
    name_to_id: dict[str, int] = {}
    pat = re.compile(r"\| (\d+) \| `([^`]+)` \|")
    for m in pat.finditer(fault_md):
        fid, name = int(m.group(1)), m.group(2).strip()
        name_to_id[name] = fid
        name_to_id["RBMS_" + name] = fid
    return name_to_id


def _lookup_fid(name_to_id: dict[str, int], map_sig: str) -> int | None:
    map_sig = str(map_sig).split(" || ")[0].strip()
    if map_sig in name_to_id:
        return name_to_id[map_sig]
    if map_sig.startswith("RBMS_"):
        return name_to_id.get(map_sig[5:])
    return None


def _resolve_addr(expr: object, i: int = 0) -> int | None:
    if isinstance(expr, (int, float)):
        return int(expr)
    s = str(expr).strip()
    m = re.match(r"i \* (\d+) \+ (\d+)(?: \+ j)?", s)
    return int(m.group(1)) * i + int(m.group(2)) if m else None


def _array_base(expr: object) -> int | None:
    m = re.match(r"i \* (\d+) \+ (\d+) \+ j", str(expr).strip())
    return int(m.group(2)) if m else None


DiscreteMap = dict[int, tuple[int, int | None]]


def _load_discrete_maps() -> tuple[DiscreteMap, DiscreteMap]:
    """Return (bank_map, rack_template@i=0) addr -> (fid, expect)."""
    name_to_id = _load_name_to_fid()
    fault_bit = _load_fault_bits()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["离散输入寄存器"]
    bank_map: dict[int, tuple[int, int | None]] = {}
    rack_map: dict[int, tuple[int, int | None]] = {}
    for r in range(2, ws.max_row + 1):
        layer = ws.cell(r, 9).value
        src = ws.cell(r, 11).value
        map_sig = ws.cell(r, 12).value
        field = ws.cell(r, 1).value
        addr_expr = ws.cell(r, 2).value
        if _is_reserved_row(field, map_sig) or _is_reserved_addr_expr(addr_expr):
            continue
        if not src or "RBMS" not in str(src):
            continue
        fid_name: str | None = None
        rack_addr: int | None = None
        if layer == "Bank":
            if not map_sig:
                continue
            fid_name = str(map_sig)
        elif layer == "Rack":
            rack_addr = _resolve_addr(addr_expr, 0)
            if rack_addr is None:
                continue
            if map_sig:
                fid_name = str(map_sig)
            elif rack_addr in RACK_DISCRETE_SUPPLEMENT_RBMS_ONLY:
                fid_name = RACK_DISCRETE_SUPPLEMENT_RBMS_ONLY[rack_addr]
            else:
                continue
        else:
            continue
        fid = _lookup_fid(name_to_id, fid_name)
        if fid is None:
            continue
        if layer == "Bank" and isinstance(addr_expr, (int, float)):
            bank_map[int(addr_expr)] = (fid, fault_bit.get(fid))
        elif layer == "Rack" and rack_addr is not None:
            rack_map[rack_addr] = (fid, fault_bit.get(fid))
    wb.close()
    return bank_map, rack_map


def _raw_db(profile_name: str) -> dict[str, RawSignal]:
    p = get_profile(profile_name)
    s = load_matrix_csv(p.default_csv, skip_signals=p.skip_signals)
    out: dict[str, RawSignal] = {}
    for sig in s.signals:
        raw = physical_to_raw(sig.value, sig.resolution, sig.offset)
        is_u32 = sig.bit_len > 16 or sig.data_type.lower() == "uint32"
        out[sig.signal] = RawSignal(raw=raw, is_u32=is_u32)
    return out


def _split_u32(raw: int) -> tuple[int, int]:
    raw &= 0xFFFFFFFF
    return raw & 0xFFFF, (raw >> 16) & 0xFFFF


def _signals_by_prefix(db: dict[str, RawSignal], prefix: str) -> list[tuple[int, RawSignal]]:
    pat = re.compile(rf"^{re.escape(prefix)}_(\d+)$")
    items: list[tuple[int, RawSignal]] = []
    for name, rs in db.items():
        m = pat.match(name)
        if m:
            items.append((int(m.group(1)), rs))
    items.sort(key=lambda x: x[0])
    return items


def _build_input_expected() -> dict[int, int]:
    dbs = {name: _raw_db(name) for name in ("suminfo", "volt", "temp", "soxdebug1", "soxdebug2")}
    expected: dict[int, int] = {}

    def put_scalar(offset: int, profile: str, signal: str) -> None:
        rs = dbs[profile][signal]
        if rs.is_u32:
            lo, hi = _split_u32(rs.raw)
            expected[offset] = lo
            expected[offset + 1] = hi
        else:
            expected[offset] = rs.raw

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["输入寄存器"]
    for r in range(2, ws.max_row + 1):
        layer = ws.cell(r, 9).value
        src = str(ws.cell(r, 11).value or "")
        if layer != "Rack" or "RBMS" not in src:
            continue
        addr_expr = ws.cell(r, 2).value
        map_sig = ws.cell(r, 12).value
        field = ws.cell(r, 1).value
        if _is_reserved_row(field, map_sig) or _is_reserved_addr_expr(addr_expr):
            continue
        base = _array_base(addr_expr)
        if base is not None:
            profile = PROFILE_BY_SRC[src]
            prefix = str(map_sig).strip() if map_sig else ""
            if not prefix:
                continue
            for idx, rs in _signals_by_prefix(dbs[profile], prefix):
                expected[base + idx - 1] = rs.raw
            continue
        offset = _resolve_addr(addr_expr, 0)
        if offset is None:
            continue
        if map_sig:
            profile = PROFILE_BY_SRC[src]
            put_scalar(offset, profile, str(map_sig).strip())
        elif offset in SOX_BY_OFFSET:
            spec = SOX_BY_OFFSET[offset]
            if spec is None:
                expected[offset] = 0x001F if offset == 2238 else 0x003F
            else:
                rs = dbs[spec[0]][spec[1]]
                if rs.is_u32:
                    lo, hi = _split_u32(rs.raw)
                    expected[offset] = lo
                    expected[offset + 1] = hi
                else:
                    expected[offset] = rs.raw
    wb.close()
    return expected


def _read_discrete(client: ModbusTcpClient, slave: int, start: int, count: int) -> list[int]:
    rr = client.read_discrete_inputs(start, count=count, device_id=slave)
    if rr.isError():
        msg = f"read_discrete_inputs({start},{count}) failed: {rr}"
        raise RuntimeError(msg)
    return [int(b) for b in rr.bits[:count]]


def _read_input_regs(client: ModbusTcpClient, slave: int, start: int, count: int) -> list[int]:
    rr = client.read_input_registers(start, count=count, device_id=slave)
    if rr.isError():
        msg = f"read_input_registers({start},{count}) failed: {rr}"
        raise RuntimeError(msg)
    return [v if v < 32768 else v - 65536 for v in rr.registers]


def _read_input_regs_chunked(
    client: ModbusTcpClient,
    slave: int,
    segments: list[tuple[int, int]],
) -> dict[int, int]:
    poll: dict[int, int] = {}
    for start, count in segments:
        remaining = count
        pos = start
        while remaining > 0:
            chunk = min(remaining, MAX_REGS_PER_READ)
            vals = _read_input_regs(client, slave, pos, chunk)
            for i, v in enumerate(vals):
                poll[pos + i] = v
            pos += chunk
            remaining -= chunk
    return poll


def _compare_discrete(
    label: str,
    addr_map: dict[int, tuple[int, int | None]],
    poll: dict[int, int],
    addr_range: range,
) -> tuple[int, int]:
    ok = bad = 0
    mismatch: list[tuple] = []
    for addr in addr_range:
        if addr not in addr_map:
            continue
        fid, exp = addr_map[addr]
        got = poll.get(addr, -1)
        if got == exp:
            ok += 1
        else:
            bad += 1
            mismatch.append((addr, got, exp, fid))
    print(f"\n=== {label} ===")
    print(f"映射点: {ok + bad}  一致: {ok}  不符: {bad}")
    for addr, got, exp, fid in mismatch[:20]:
        print(f"  addr={addr} poll={got} expect={exp} FID={fid}")
    if len(mismatch) > 20:
        print(f"  ... 另有 {len(mismatch) - 20} 条不符")
    return ok, bad


def _compare_input(
    label: str,
    expected: dict[int, int],
    poll: dict[int, int],
) -> tuple[int, int]:
    ok = bad = 0
    mismatch: list[tuple] = []
    for addr in sorted(expected):
        exp_u = expected[addr] & 0xFFFF
        got = poll.get(addr)
        if got is None:
            bad += 1
            mismatch.append((addr, None, exp_u))
            continue
        got_u = got & 0xFFFF
        if got_u == exp_u:
            ok += 1
        else:
            bad += 1
            mismatch.append((addr, got_u, exp_u))
    print(f"\n=== {label} ===")
    print(f"映射点: {ok + bad}  一致: {ok}  不符: {bad}")
    for addr, got, exp in mismatch[:20]:
        print(f"  addr={addr} poll={got} expect={exp}")
    if len(mismatch) > 20:
        print(f"  ... 另有 {len(mismatch) - 20} 条不符")
    return ok, bad


def _find_working_slave(host: str, port: int, slaves: range) -> int:
    for sid in slaves:
        client = ModbusTcpClient(host=host, port=port, timeout=3)
        if not client.connect():
            continue
        try:
            rr = client.read_discrete_inputs(200, count=1, device_id=sid)
            if not rr.isError():
                print(f"使用 Slave ID = {sid}")
                client.close()
                return sid
        except Exception:
            pass
        client.close()
    msg = f"无法在 {host}:{port} 上找到可用 Slave ID（尝试 {slaves.start}-{slaves.stop - 1}）"
    raise RuntimeError(msg)


def _rack_discrete_addr(template_addr: int, rack_idx: int) -> int:
    return template_addr + rack_idx * DISCRETE_RACK_STRIDE


def _rack_input_addr(offset: int, rack_idx: int) -> int:
    return rack_idx * INPUT_RACK_STRIDE + offset


def main() -> int:
    parser = argparse.ArgumentParser(description="Verify Modbus vs rbms_tcp_sim config")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--slave", type=int, default=0, help="0=auto scan 1-247")
    parser.add_argument("--racks", type=int, default=4, help="核对簇数 1-4")
    args = parser.parse_args()
    racks = max(1, min(4, args.racks))

    if args.slave > 0:
        slave = args.slave
    else:
        slave = _find_working_slave(args.host, args.port, range(1, 248))

    bank_map, rack_template = _load_discrete_maps()
    input_expected = _build_input_expected()

    client = ModbusTcpClient(host=args.host, port=args.port, timeout=5)
    if not client.connect():
        print(f"连接失败: {args.host}:{args.port}")
        return 2
    print(f"已连接 Modbus TCP {args.host}:{args.port} slave={slave} racks={racks}")

    total_bad = 0
    try:
        di_bank = _read_discrete(client, slave, 1, 125)
        poll_di = {i + 1: di_bank[i] for i in range(125)}

        _, b_bad = _compare_discrete("离散输入 堆级 1-125", bank_map, poll_di, range(1, 126))
        total_bad += b_bad

        for rack_idx in range(racks):
            di_start = _rack_discrete_addr(200, rack_idx)
            di_vals = _read_discrete(client, slave, di_start, 200)
            rack_poll = {di_start + i: di_vals[i] for i in range(200)}
            rack_map = {
                _rack_discrete_addr(addr, rack_idx): val for addr, val in rack_template.items()
            }
            _, r_bad = _compare_discrete(
                f"离散输入 簇{rack_idx + 1} {di_start}-{di_start + 199}",
                rack_map,
                rack_poll,
                range(di_start, di_start + 200),
            )
            total_bad += r_bad

        for rack_idx in range(racks):
            base = rack_idx * INPUT_RACK_STRIDE
            segments = [(base + off, cnt) for off, cnt, _ in INPUT_SEGMENTS]
            poll_ir = _read_input_regs_chunked(client, slave, segments)
            rack_expected = {
                _rack_input_addr(off, rack_idx): raw for off, raw in input_expected.items()
            }
            for seg_off, seg_cnt, seg_label in INPUT_SEGMENTS:
                seg_start = base + seg_off
                seg_addrs = range(seg_start, seg_start + seg_cnt)
                seg_expected = {a: rack_expected[a] for a in seg_addrs if a in rack_expected}
                if not seg_expected:
                    continue
                _, ir_bad = _compare_input(
                    f"输入寄存器 簇{rack_idx + 1} {seg_label} "
                    f"({seg_addrs.start}-{seg_addrs.stop - 1})",
                    seg_expected,
                    poll_ir,
                )
                total_bad += ir_bad

        print(f"\n总计不符: {total_bad}")
        return 0 if total_bad == 0 else 1
    finally:
        client.close()


if __name__ == "__main__":
    raise SystemExit(main())
