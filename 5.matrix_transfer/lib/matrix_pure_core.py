#!/usr/bin/env python3
"""Shared pure-Matrix point-table helpers for RBMS/BBMS generators."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from lib.matrix_point_gen import (
    POINTATTR_MATRIX_SOURCE_COMMENT,
    MatrixSignal,
    MergedPointAttr,
    precision_from_ratio,
    wrap_clang_format_off,
    _compute_pointattr_layout,
    format_pointattr_row,
)

MATRIX_BYTE_BASE = 1

MATRIX_MACRO_VALUES: dict[str, int] = {
    "TOTAL_SERIES_CELL_NUMBER_PER_RACK_BMS": 416,
    "TOTAL_SERIES_CELL_NUMBER_PER_RACK_MAX_REAL": 416,
    "RACK_TEMPERATURE_SENSOR_MUX_NUMBER": 64,
    "RACK_AFE_NUMBER_MAX": 32,
}

_UNIT_PAREN_CONTENTS: frozenset[str] = frozenset(
    {
        "%",
        "℃",
        "°c",
        "v",
        "a",
        "mv",
        "kv",
        "kw",
        "kwh",
        "kohm",
        "ohm",
        "min",
        "day",
        "bar",
        "ma",
        "mah",
        "ah",
        "flg",
        "高精度",
        "低精度",
        "high accu",
        "low accu",
    }
)


@dataclass
class MatrixRow:
    signal: MatrixSignal
    byte_raw: object
    row_index: int


@dataclass
class PointSpec:
    enum_name: str
    tail_enum_name: str
    name_zh: str
    name_en: str
    message: str
    cmd_group: int
    cmd_id: int
    section_order: int
    member_order: float
    data_idx: int
    data_start_bit: int
    data_bit_len: int
    data_type: str
    coeff: float
    offset: float
    max_val: float
    min_val: float
    unit: str
    precision: int
    repeat_cnt: int
    matrix_signal: str
    group_type: int = 1


@dataclass
class EnumEmitRow:
    enum_name: str
    comment: str
    assign_expr: str | None = None
    section_comment: str = ""
    code_slots: int = 1
    is_fold_tail: bool = False


@dataclass
class CsvEmitRow:
    name_zh: str
    name_en: str
    code: int
    spec: PointSpec | None
    group_type: int = 1


@dataclass
class PointAttrEmitRow:
    point_id: str
    data_idx: int
    data_bit_len: int
    data_start_bit: int
    data_type: str
    coeff: float
    offset: float
    max_val: float
    min_val: float
    repeat_cnt: int


@dataclass
class SectionAnchor:
    order: int
    section_comment: str
    matrix_messages: tuple[str, ...]
    cmd_group: int
    cmd_id: int
    pointattr_array: str
    pointattr_title: str
    cmd_id_notes: tuple[str, ...] = ()


@dataclass
class MessageBlock:
    anchor: SectionAnchor
    specs: list[PointSpec] = field(default_factory=list)
    pointattr_rows: list[PointAttrEmitRow] = field(default_factory=list)
    total_bytes: int = 0


def format_matrix_cmdid_comment(cmd_id: int) -> str:
    return f"{cmd_id} (0x{cmd_id:02X})"


def format_matrix_cmd_group_comment(cmd_group: int) -> str:
    return f"0x{cmd_group:02X}"


def parse_matrix_description(description: str, signal_name: str) -> tuple[str, str, bool]:
    lines = [line.strip() for line in (description or "").splitlines() if line.strip()]
    if len(lines) >= 2:
        return lines[1], lines[0], True
    if len(lines) == 1:
        text = lines[0]
        if re.search(r"[\u4e00-\u9fff]", text):
            return text, signal_name, True
        return text, text, False
    return signal_name, signal_name, False


def _paren_content_is_unit_like(inner: str) -> bool:
    text = inner.strip()
    if not text:
        return False
    if text in ("%", "℃"):
        return True
    lowered = text.lower()
    if lowered in _UNIT_PAREN_CONTENTS:
        return True
    if re.fullmatch(r"[a-zA-Z%℃°]+", text):
        return True
    return False


def _paren_content_is_array_capacity_hint(inner: str) -> bool:
    text = inner.strip()
    if re.fullmatch(r"最多\d+个", text):
        return True
    return bool(re.fullmatch(r"up to \d+ .+", text, flags=re.IGNORECASE))


def strip_trailing_unit_suffix(text: str) -> str:
    def replace(match: re.Match[str]) -> str:
        inner = match.group(1)
        if _paren_content_is_unit_like(inner) or _paren_content_is_array_capacity_hint(inner):
            return ""
        return match.group(0)

    cleaned = re.sub(r"\s*[(（]([^）)]+)[)）]", replace, text)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return cleaned.strip()


def csv_display_name(text: str) -> str:
    return strip_trailing_unit_suffix(text)


def _to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_byte_field(byte_value: object) -> tuple[int, int] | None:
    if byte_value is None or byte_value == "":
        return None
    text = str(byte_value).strip()
    if "-" in text:
        left, right = text.split("-", 1)
        return int(left.strip()), int(right.strip())
    value = int(text)
    return value, value


def byte_span_count(byte_value: object) -> int:
    parsed = parse_byte_field(byte_value)
    if parsed is None:
        return 0
    low, high = parsed
    return high - low + 1


def eval_bracket_size(expr: str) -> int | None:
    expr = expr.strip()
    if expr.isdigit():
        return int(expr)
    if expr in MATRIX_MACRO_VALUES:
        return MATRIX_MACRO_VALUES[expr]
    if "*" in expr:
        left, right = expr.split("*", 1)
        left_val = eval_bracket_size(left.strip())
        right_val = eval_bracket_size(right.strip())
        if left_val is not None and right_val is not None:
            return left_val * right_val
    if "/" in expr:
        left, right = expr.split("/", 1)
        left_val = eval_bracket_size(left.strip())
        right_val = eval_bracket_size(right.strip())
        if left_val is not None and right_val is not None and right_val != 0:
            return left_val // right_val
    return None


def bracket_repeat_hint(signal_name: str) -> int | None:
    match = re.search(r"\[([^\]]+)\]", signal_name)
    if not match:
        return None
    inner = match.group(1)
    if inner.isdigit():
        bits = int(inner)
        return (bits + 7) // 8 if bits > 32 else bits
    return eval_bracket_size(inner)


def is_matrix_array_signal(signal_name: str) -> bool:
    return "[" in signal_name


def is_fault_bit_array_signal(signal_name: str) -> bool:
    return bool(re.match(r"^(RBMS|BBMS(_A)?)_Fault\[", signal_name))


def enum_name_contains_reserved(enum_name: str) -> bool:
    return "reserved" in enum_name.lower()


def disambiguate_reserved_enum(enum_name: str, cmd_id: int) -> str:
    if not enum_name_contains_reserved(enum_name):
        return enum_name
    if re.search(r"_[0-9A-Fa-f]{2}$", enum_name):
        return enum_name
    return f"{enum_name}_{cmd_id:02d}"


def compute_repeat_cnt(signal: MatrixSignal, byte_raw: object) -> int:
    span = byte_span_count(byte_raw)
    if span <= 0:
        return 0
    if not is_matrix_array_signal(signal.signal_name):
        return 0
    bit_len = signal.bit_len or 0
    if bit_len == 1:
        return span
    if bit_len == 8:
        return span
    if bit_len == 16:
        return span * 8 // 16
    if bit_len == 32:
        return span * 8 // 32
    bracket = bracket_repeat_hint(signal.signal_name)
    if bracket is not None and bracket > 1:
        return bracket
    return 0


def pointattr_bit_len(signal: MatrixSignal, repeat_cnt: int) -> int:
    raw = signal.bit_len or 8
    if raw == 1 and repeat_cnt > 1:
        return 8
    return raw


def infer_data_type(signal: MatrixSignal, point_bit_len: int, repeat_cnt: int) -> str:
    del signal, repeat_cnt
    if point_bit_len > 16:
        return "Uint32"
    if point_bit_len > 8:
        return "Uint16"
    return "Uint8"


def normalize_max_val(max_val: float, repeat_cnt: int, bit_len: int) -> float:
    if repeat_cnt > 1 and bit_len == 8 and max_val <= 1.0:
        return 255.0
    return max_val


def array_enum_base_name(enum_name: str, signal_name: str) -> str:
    base = enum_name
    if "CellSdrate" in base:
        base = re.sub(r"_\d+$", "", base)
    elif "CellBalStatus" in base:
        base = re.sub(r"1_8$", "", base)
    elif "_Byte_" in base:
        base = re.sub(r"_Byte_\d+$", "", base)
    elif is_fault_bit_array_signal(signal_name):
        base = re.sub(r"_Byte_\d+$", "", base)
        if base.endswith("_Byte"):
            base = base[: -len("_Byte")]
    elif "SaSGPC_CellVmVxT" in base:
        base = re.sub(r"_\d+$", "", base)
    elif re.search(r"Fault", base):
        base = re.sub(r"_\d+$", "", base)
    return base


def fold_head_enum_name(enum_name: str, signal_name: str, repeat_cnt: int) -> str:
    if repeat_cnt <= 1:
        return enum_name
    return f"{array_enum_base_name(enum_name, signal_name)}_Start"


def fold_tail_enum_name(start_enum: str, repeat_cnt: int) -> str:
    if repeat_cnt <= 1:
        return start_enum
    if start_enum.endswith("_Start"):
        return f"{start_enum[: -len('_Start')]}_End"
    return f"{start_enum}_End"


def load_comm_matrix_by_message(matrix_path: Path) -> dict[str, list[MatrixRow]]:
    from openpyxl import load_workbook

    workbook = load_workbook(matrix_path, read_only=True, data_only=True)
    sheet = workbook["Comm Matrix"]
    grouped: dict[str, list[MatrixRow]] = {}
    current_message = ""
    row_index = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_index += 1
        message_name = row[0]
        signal_name = row[2]
        if message_name:
            current_message = str(message_name).strip()
        if not signal_name or not current_message:
            continue
        signal_key = str(signal_name).strip()
        byte_raw = row[4]
        byte_parsed = parse_byte_field(byte_raw)
        unit = str(row[12] or "").strip()
        if unit in ("/", "None", "none"):
            unit = ""
        signal = MatrixSignal(
            signal_name=signal_key,
            message_name=current_message,
            description=str(row[3] or "").strip(),
            byte=byte_parsed[0] if byte_parsed else _to_int(byte_raw),
            start_bit=_to_int(row[5]),
            bit_len=_to_int(row[6]),
            resolution=_to_float(row[7]),
            offset=_to_float(row[8]),
            min_val=_to_float(row[9]),
            max_val=_to_float(row[10]),
            unit=unit,
        )
        grouped.setdefault(current_message, []).append(
            MatrixRow(signal=signal, byte_raw=byte_raw, row_index=row_index)
        )
    return grouped


def load_matrix_message_descriptions(matrix_path: Path) -> dict[str, str]:
    from openpyxl import load_workbook

    workbook = load_workbook(matrix_path, read_only=True, data_only=True)
    sheet = workbook["Comm Matrix"]
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        message_name = row[0]
        message_desc = row[1]
        if not message_name:
            continue
        name = str(message_name).strip()
        if message_desc:
            result[name] = str(message_desc).strip().replace("\n", " ")
        elif name not in result:
            result[name] = ""
    return result


def build_enum_rows(
    blocks: list[MessageBlock],
    *,
    online_enum: str,
    online_zh: str,
    online_section: str,
    data_end_enum: str,
    data_end_zh: str,
) -> list[EnumEmitRow]:
    rows: list[EnumEmitRow] = [
        EnumEmitRow(
            enum_name=online_enum,
            comment=online_zh,
            assign_expr="0",
            section_comment=online_section,
            code_slots=1,
        )
    ]
    last_section = ""
    pending_assign: str | None = None
    for block in blocks:
        if block.anchor.section_comment != last_section:
            rows.append(
                EnumEmitRow(
                    enum_name="",
                    comment="",
                    section_comment=block.anchor.section_comment,
                )
            )
            last_section = block.anchor.section_comment
        for spec in block.specs:
            if spec.repeat_cnt > 1:
                rows.append(
                    EnumEmitRow(
                        enum_name=spec.enum_name,
                        comment=spec.name_zh,
                        assign_expr=pending_assign,
                        code_slots=1,
                    )
                )
                rows.append(
                    EnumEmitRow(
                        enum_name=spec.tail_enum_name,
                        comment=spec.name_zh,
                        assign_expr=f"{spec.enum_name} + {spec.repeat_cnt}",
                        code_slots=spec.repeat_cnt - 1,
                        is_fold_tail=True,
                    )
                )
                pending_assign = spec.tail_enum_name
            else:
                rows.append(
                    EnumEmitRow(
                        enum_name=spec.enum_name,
                        comment=spec.name_zh,
                        assign_expr=pending_assign,
                        code_slots=1,
                    )
                )
                pending_assign = None
    # Data_End: no explicit assign — C enum auto-increments to last point Code + 1 (RTDB count).
    rows.append(
        EnumEmitRow(
            enum_name=data_end_enum,
            comment=data_end_zh,
            assign_expr=None,
            code_slots=1,
        )
    )
    return rows


def expand_csv_rows(
    blocks: list[MessageBlock],
    enum_rows: list[EnumEmitRow],
    *,
    online_zh: str,
    online_en: str,
) -> list[CsvEmitRow]:
    code_by_enum: dict[str, int] = {}
    code = 0
    for row in enum_rows:
        if not row.enum_name:
            continue
        code_by_enum[row.enum_name] = code
        code += row.code_slots

    csv_rows: list[CsvEmitRow] = [
        CsvEmitRow(name_zh=online_zh, name_en=online_en, code=0, spec=None, group_type=0)
    ]
    for block in blocks:
        for spec in block.specs:
            if spec.repeat_cnt > 1:
                zh_base = csv_display_name(spec.name_zh)
                en_base = csv_display_name(spec.name_en)
                if "[" in zh_base:
                    zh_base = re.sub(r"\[\d+\]", "", zh_base)
                if "[" in en_base:
                    en_base = re.sub(r"\[\d+\]", "", en_base)
                start_code = code_by_enum.get(spec.enum_name, len(csv_rows))
                for index in range(spec.repeat_cnt):
                    csv_rows.append(
                        CsvEmitRow(
                            name_zh=f"{zh_base}[{index}]",
                            name_en=f"{en_base} [{index}]",
                            code=start_code + index,
                            spec=spec,
                        )
                    )
            else:
                start_code = code_by_enum.get(spec.enum_name, len(csv_rows))
                csv_rows.append(
                    CsvEmitRow(
                        name_zh=csv_display_name(spec.name_zh),
                        name_en=csv_display_name(spec.name_en),
                        code=start_code,
                        spec=spec,
                    )
                )
    return csv_rows


def render_enum_snippet(
    enum_rows: list[EnumEmitRow],
    *,
    header_comment: str,
    typedef_suffix: str,
) -> str:
    enum_pairs: list[tuple[str, str]] = []
    prev_was_content = False
    for row in enum_rows:
        if row.section_comment and not row.enum_name:
            if prev_was_content:
                enum_pairs.append(("__BLANK__", ""))
            enum_pairs.append(("__MSGHDR__", row.section_comment))
            prev_was_content = True
            continue
        if not row.enum_name:
            continue
        comment = row.comment if row.comment else row.enum_name
        if row.assign_expr:
            enum_pairs.append((f"{row.enum_name} = {row.assign_expr}", comment))
        else:
            enum_pairs.append((row.enum_name, comment))
        prev_was_content = True

    aligned = [(code, cmt) for code, cmt in enum_pairs if code and code not in ("__BLANK__", "__MSGHDR__")]
    max_code_len = max((len(code) for code, _ in aligned), default=0)
    lines = [
        header_comment,
        "typedef enum",
        "{",
    ]
    for code, comment in enum_pairs:
        if code == "__BLANK__":
            lines.append("")
            continue
        if code == "__MSGHDR__":
            lines.append(f"    //{comment}")
            continue
        pad = max(1, max_code_len - len(code) + 1)
        lines.append(f"    {code},{' ' * pad}// {comment}")
    lines.append(f"}} {typedef_suffix};")
    return wrap_clang_format_off("\n".join(lines))


def pointattr_row_to_merged(row: PointAttrEmitRow, array_name: str) -> MergedPointAttr:
    return MergedPointAttr(
        point_id=row.point_id,
        array_name=array_name,
        data_idx=row.data_idx,
        data_bit_len=row.data_bit_len,
        data_start_bit=row.data_start_bit,
        data_type=row.data_type,
        coeff=row.coeff,
        offset=row.offset,
        max_val=row.max_val,
        min_val=row.min_val,
        repeat_cnt=row.repeat_cnt,
    )


def render_pointattr_snippet(
    blocks: list[MessageBlock],
    descriptions: dict[str, str],
    *,
    header_lines: tuple[str, ...],
    rbms_style: bool,
) -> str:
    lines = list(header_lines) + [POINTATTR_MATRIX_SOURCE_COMMENT, ""]
    for block in blocks:
        title = block.anchor.pointattr_title
        msg = block.anchor.matrix_messages[0] if block.anchor.matrix_messages else ""
        desc = descriptions.get(msg, "")
        if desc and desc not in title:
            title = f"{title} ({block.total_bytes} Bytes)"
        cmd_notes = ""
        if block.anchor.cmd_id_notes:
            cmd_notes = f"; also: {', '.join(block.anchor.cmd_id_notes)}"
        lines.extend(
            [
                "/*============================================================",
                f" * {title}",
                f" * 对应消息ID: {msg} (cmdGroup: {format_matrix_cmd_group_comment(block.anchor.cmd_group)}, "
                f"cmdid: {format_matrix_cmdid_comment(block.anchor.cmd_id)}{cmd_notes})",
                " *============================================================*/",
                f"const bmsPointAttr_t {block.anchor.pointattr_array} =",
                "{",
                "    /* pointId,                dataIdx,  dataBitLen,  dataStartBit,  dataType,  coeff,    offset,    maxVal,          minVal,   repeatCnt */",
            ]
        )
        entries = [
            pointattr_row_to_merged(row, block.anchor.pointattr_array)
            for row in block.pointattr_rows
        ]
        layout = _compute_pointattr_layout(entries, rbms_style=rbms_style)
        for entry in entries:
            lines.append(format_pointattr_row(entry, layout))
        lines.append("};")
        lines.append("")
    return wrap_clang_format_off("\n".join(lines))


def render_csv_dict_rows(csv_rows: list[CsvEmitRow]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in csv_rows:
        spec = item.spec
        ratio = spec.coeff if spec else 1.0
        offset = spec.offset if spec else 0.0
        precision = spec.precision if spec else 0
        max_val = "" if spec is None else (str(spec.max_val) if spec.max_val or spec.max_val == 0 else "")
        min_val = "" if spec is None else (str(spec.min_val) if spec.min_val or spec.min_val == 0 else "")
        unit = spec.unit if spec else ""
        rows.append(
            {
                "Name": item.name_zh,
                "Ename": item.name_en,
                "Code": str(item.code),
                "Group Type": str(item.group_type),
                "Attribute": "0",
                "Function Code": "0",
                "Data Type": "0",
                "Register Address": "0",
                "Bit Position": "0",
                "Bit Number": "0",
                "Precision": str(precision),
                "Ratio": str(ratio),
                "Offset": str(offset),
                "Endian": "0",
                "Is Persisted": "1",
                "Storage Interval": "30000",
                "Mutate Bound": "0",
                "Default Value": "",
                "Max Value": max_val,
                "Min Value": min_val,
                "Unit": unit,
                "Is Show": "0",
            }
        )
    return rows
