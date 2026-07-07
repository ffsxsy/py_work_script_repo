#!/usr/bin/env python3
"""Pure-Matrix BBMS point table generator (see docs/design/BBMS_Matrix_纯规范生成计划.md)."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from lib.matrix_point_gen import GenReport, precision_from_ratio, write_csv, write_text
from lib.matrix_pure_core import (
    MATRIX_BYTE_BASE,
    MatrixRow,
    MessageBlock,
    PointAttrEmitRow,
    PointSpec,
    SectionAnchor,
    build_enum_rows,
    compute_repeat_cnt,
    disambiguate_reserved_enum,
    expand_csv_rows,
    fold_head_enum_name,
    fold_tail_enum_name,
    infer_data_type,
    load_comm_matrix_by_message,
    load_matrix_message_descriptions,
    normalize_max_val,
    parse_byte_field,
    parse_matrix_description,
    pointattr_bit_len,
    render_csv_dict_rows,
    render_enum_snippet,
    render_pointattr_snippet,
)

ONLINE_ENUM = "kBbms_Online"
ONLINE_ZH = "在线状态"
ONLINE_EN = "Online State"
DATA_END_ENUM = "kBbms_Data_End"
DATA_END_ZH = "BBMS测点结束"

CANONICAL_BBMSNO_MESSAGE = "BBMS_SumInfo"
HMI_BBMSNO_MESSAGES = frozenset(
    {
        "HMI_CtlWord",
        "HMI_BankFaultCali",
        "HMI_FltOvTiNbr",
    }
)


@dataclass(frozen=True)
class MessageIdRow:
    message_name: str
    cmd_group: int
    cmd_id: int
    cmd_type: str
    req_payload: str
    resp_payload: str
    hmi: str
    bbms_a: str
    bbms_m: str
    rbms: str
    scope_reason: str


@dataclass
class ScopedMessage:
    message_name: str
    cmd_group: int
    cmd_id: int
    reasons: frozenset[str]
    cmd_pairs: tuple[tuple[int, int], ...]


def _norm_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int_field(value: object) -> int:
    text = _norm_cell(value)
    if not text:
        return 0
    if text.lower().startswith("0x"):
        return int(text, 16)
    return int(float(text))


def _is_nonempty_payload(value: object) -> bool:
    text = _norm_cell(value)
    return bool(text) and text not in ("/", "-", "None", "none")


def _is_full_struct_response(value: object) -> bool:
    text = _norm_cell(value)
    if not text or text == "/":
        return False
    lowered = text.lower()
    if "struct state" in lowered:
        return False
    return "struct" in lowered


def _is_hmi_rbms_forward(row: MessageIdRow) -> bool:
    return (
        row.message_name.startswith("HMI_")
        and row.cmd_group == 0x03
        and row.rbms == "Tx"
    )


def load_message_id_rows(matrix_path: Path) -> list[MessageIdRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(matrix_path, read_only=True, data_only=True)
    sheet = workbook["Message ID"]
    rows: list[MessageIdRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        message_name = _norm_cell(row[0])
        if not message_name:
            continue
        rows.append(
            MessageIdRow(
                message_name=message_name,
                cmd_group=_parse_int_field(row[4]),
                cmd_id=_parse_int_field(row[5]),
                cmd_type=_norm_cell(row[6]),
                req_payload=_norm_cell(row[8]),
                resp_payload=_norm_cell(row[9]),
                hmi=_norm_cell(row[11]),
                bbms_a=_norm_cell(row[12]),
                bbms_m=_norm_cell(row[13]),
                rbms=_norm_cell(row[14]),
                scope_reason="",
            )
        )
    return rows


def _classify_scope(row: MessageIdRow) -> str | None:
    if row.message_name.startswith("RBMS_"):
        return None
    if row.bbms_m == "Tx" or row.bbms_a == "Tx":
        if _is_nonempty_payload(row.req_payload):
            return "TX"
    cmd_type = row.cmd_type.lower()
    if cmd_type in ("0x02", "2"):
        if _is_full_struct_response(row.resp_payload):
            return "RD"
        if row.hmi == "Tx" and (row.bbms_m == "Rx" or row.bbms_a == "Rx"):
            if _is_nonempty_payload(row.req_payload):
                return "WR"
    return None


def filter_bbms_scope(rows: list[MessageIdRow]) -> list[ScopedMessage]:
    grouped: dict[str, list[tuple[MessageIdRow, str]]] = {}
    for row in rows:
        reason = _classify_scope(row)
        if reason is None:
            continue
        if _is_hmi_rbms_forward(row):
            continue
        grouped.setdefault(row.message_name, []).append((row, reason))

    scoped: list[ScopedMessage] = []
    for message_name, entries in grouped.items():
        reasons = frozenset(reason for _, reason in entries)
        cmd_pairs = tuple(sorted({(item.cmd_group, item.cmd_id) for item, _ in entries}))
        cmd_group, cmd_id = cmd_pairs[0]
        scoped.append(
            ScopedMessage(
                message_name=message_name,
                cmd_group=cmd_group,
                cmd_id=cmd_id,
                reasons=reasons,
                cmd_pairs=cmd_pairs,
            )
        )
    scoped.sort(key=lambda item: (item.cmd_group, item.cmd_id, item.message_name))
    return scoped


def build_bbms_section_order(
    scoped: list[ScopedMessage],
    descriptions: dict[str, str],
) -> list[SectionAnchor]:
    anchors: list[SectionAnchor] = [
        SectionAnchor(0, "BBMS设备模型", (), 0, 0, "", ""),
    ]
    for index, item in enumerate(scoped, start=1):
        title = descriptions.get(item.message_name, item.message_name)
        if len(title) > 80:
            title = item.message_name
        extra = [
            f"0x{grp:02X}:{cid}"
            for grp, cid in item.cmd_pairs
            if (grp, cid) != (item.cmd_group, item.cmd_id)
        ]
        anchors.append(
            SectionAnchor(
                order=index,
                section_comment=title,
                matrix_messages=(item.message_name,),
                cmd_group=item.cmd_group,
                cmd_id=item.cmd_id,
                pointattr_array=(
                    f"bbmsCmd{item.cmd_group:02d}_{item.cmd_id:02d}_{item.message_name}_PointAttr[]"
                ),
                pointattr_title=f"{item.message_name} - {title}",
                cmd_id_notes=tuple(extra),
            )
        )
    return anchors


def matrix_signal_to_bbms_enum(signal_name: str) -> str:
    base = re.sub(r"\[.*\]", "", signal_name).strip()
    if base.startswith("BBMS_"):
        return "kBbms_" + base[5:]
    return "kBbms_" + base


def should_skip_bbmsno_signal(message_name: str, signal_name: str) -> bool:
    if signal_name != "BBMSNo":
        return False
    if message_name == CANONICAL_BBMSNO_MESSAGE:
        return False
    if message_name in HMI_BBMSNO_MESSAGES:
        return False
    return True


def matrix_to_spec(
    row: MatrixRow,
    anchor: SectionAnchor,
    report: GenReport,
) -> PointSpec | None:
    signal = row.signal
    if should_skip_bbmsno_signal(signal.message_name, signal.signal_name):
        report.info(f"跳过重复 BBMSNo: {signal.message_name}.{signal.signal_name}（canonical: {CANONICAL_BBMSNO_MESSAGE}）")
        return None
    enum_name = matrix_signal_to_bbms_enum(signal.signal_name)
    name_zh, name_en, has_explicit_zh = parse_matrix_description(signal.description, signal.signal_name)
    if not has_explicit_zh and not re.search(r"[\u4e00-\u9fff]", name_zh):
        report.error(f"{signal.signal_name}: 缺少中文 Description")
    repeat_cnt = compute_repeat_cnt(signal, row.byte_raw)
    if "[" in signal.signal_name and repeat_cnt <= 0:
        report.error(f"{signal.signal_name}: 无法推导 repeatCnt")
    coeff = signal.resolution if signal.resolution is not None else 1.0
    offset = signal.offset if signal.offset is not None else 0.0
    max_val = signal.max_val if signal.max_val is not None else 0.0
    min_val = signal.min_val if signal.min_val is not None else 0.0
    byte_parsed = parse_byte_field(row.byte_raw)
    data_idx = (byte_parsed[0] if byte_parsed else MATRIX_BYTE_BASE) - MATRIX_BYTE_BASE
    data_start_bit = signal.start_bit or 0
    bit_len = pointattr_bit_len(signal, repeat_cnt)
    data_type = infer_data_type(signal, bit_len, repeat_cnt)
    max_val = normalize_max_val(max_val, repeat_cnt, bit_len)
    head_enum = fold_head_enum_name(enum_name, signal.signal_name, repeat_cnt)
    head_enum = disambiguate_reserved_enum(head_enum, anchor.cmd_id)
    tail_enum = fold_tail_enum_name(head_enum, repeat_cnt)
    if repeat_cnt > 1 and "reserved" in tail_enum.lower():
        tail_enum = disambiguate_reserved_enum(tail_enum, anchor.cmd_id)
    return PointSpec(
        enum_name=head_enum,
        tail_enum_name=tail_enum,
        name_zh=name_zh,
        name_en=name_en,
        message=signal.message_name,
        cmd_group=anchor.cmd_group,
        cmd_id=anchor.cmd_id,
        section_order=anchor.order,
        member_order=float(row.row_index),
        data_idx=data_idx,
        data_start_bit=data_start_bit,
        data_bit_len=bit_len,
        data_type=data_type,
        coeff=coeff,
        offset=offset,
        max_val=max_val,
        min_val=min_val,
        unit=signal.unit,
        precision=precision_from_ratio(coeff),
        repeat_cnt=repeat_cnt,
        matrix_signal=signal.signal_name,
        group_type=1,
    )


def build_message_blocks(
    grouped: dict[str, list[MatrixRow]],
    anchors: list[SectionAnchor],
    report: GenReport,
) -> list[MessageBlock]:
    blocks: list[MessageBlock] = []
    for anchor in anchors:
        if anchor.order == 0:
            continue
        message = anchor.matrix_messages[0]
        if message not in grouped:
            report.error(f"scoped 报文 {message} 在 Comm Matrix 中无信号")
            continue
        specs: list[PointSpec] = []
        max_byte = 0
        for matrix_row in grouped[message]:
            spec = matrix_to_spec(matrix_row, anchor, report)
            if spec is not None:
                specs.append(spec)
            parsed = parse_byte_field(matrix_row.byte_raw)
            if parsed:
                max_byte = max(max_byte, parsed[1])
        specs.sort(key=lambda item: item.member_order)
        pointattr_rows = [
            PointAttrEmitRow(
                point_id=spec.enum_name,
                data_idx=spec.data_idx,
                data_bit_len=spec.data_bit_len,
                data_start_bit=spec.data_start_bit,
                data_type=spec.data_type,
                coeff=spec.coeff,
                offset=spec.offset,
                max_val=spec.max_val,
                min_val=spec.min_val,
                repeat_cnt=spec.repeat_cnt,
            )
            for spec in specs
        ]
        blocks.append(
            MessageBlock(
                anchor=anchor,
                specs=specs,
                pointattr_rows=pointattr_rows,
                total_bytes=max_byte,
            )
        )
    return blocks


def generate_bbms_pure(
    matrix_path: Path,
    out_dir: Path,
    report: GenReport | None = None,
) -> GenReport:
    report = report or GenReport()
    if not matrix_path.is_file():
        report.error(f"Matrix 不存在: {matrix_path}")
        return report

    message_id_rows = load_message_id_rows(matrix_path)
    scoped = filter_bbms_scope(message_id_rows)
    grouped = load_comm_matrix_by_message(matrix_path)
    descriptions = load_matrix_message_descriptions(matrix_path)
    anchors = build_bbms_section_order(scoped, descriptions)
    blocks = build_message_blocks(grouped, anchors, report)
    enum_rows = build_enum_rows(
        blocks,
        online_enum=ONLINE_ENUM,
        online_zh=ONLINE_ZH,
        online_section="BBMS设备模型",
        data_end_enum=DATA_END_ENUM,
        data_end_zh=DATA_END_ZH,
    )
    csv_emit = expand_csv_rows(blocks, enum_rows, online_zh=ONLINE_ZH, online_en=ONLINE_EN)

    bbms_dir = out_dir / "bbms"
    bbms_dir.mkdir(parents=True, exist_ok=True)
    write_text(
        bbms_dir / "devBBMSPoint_e.h.snippet",
        render_enum_snippet(
            enum_rows,
            header_comment="// BBMS 测点枚举 — generated from Matrix V1.0.50 (pure)",
            typedef_suffix="devBBMSPoint_e",
        ),
    )
    write_text(
        bbms_dir / "protocol_bms_hmi_pointattr.c.snippet",
        render_pointattr_snippet(
            blocks,
            descriptions,
            header_lines=("/* Generated bmsPointAttr_t blocks — pure Matrix BBMS */",),
            rbms_style=False,
        ),
    )
    write_csv(bbms_dir / "BBMS.csv", render_csv_dict_rows(csv_emit))
    return report
