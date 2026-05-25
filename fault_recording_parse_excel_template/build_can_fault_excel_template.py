"""Build Excel template for CAN fault recording CSV analysis.

Target: Excel 2016, 2019, 2021, Microsoft 365 (Windows, desktop).
Formulas avoid dynamic-array / LET / FILTER (365-only features).
"""

from __future__ import annotations

import importlib.util
from collections.abc import Iterator
from pathlib import Path
from types import ModuleType
from typing import Protocol, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# xlOpenXMLWorkbookMacroEnabled — same on Excel 2016 through 365
XLSM_FILE_FORMAT = 52
SUPPORTED_EXCEL_NOTE = "Excel 2016 / 2019 / 2021 / Microsoft 365 (Windows)"


def _load_sibling_module(module_name: str) -> ModuleType:
    path = Path(__file__).with_name(f"{module_name}.py")
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        msg = f"cannot load module from {path}"
        raise ImportError(msg)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


_tv = _load_sibling_module("template_version")
META_RELEASE_DATE_CELL: str = _tv.META_RELEASE_DATE_CELL
META_SUMMARY_CELL: str = _tv.META_SUMMARY_CELL
META_VERSION_CELL: str = _tv.META_VERSION_CELL
TEMPLATE_RELEASE_DATE: str = _tv.TEMPLATE_RELEASE_DATE
TEMPLATE_RELEASE_NOTES: tuple[str, ...] = _tv.TEMPLATE_RELEASE_NOTES
TEMPLATE_RELEASE_SUMMARY: str = _tv.TEMPLATE_RELEASE_SUMMARY
TEMPLATE_VERSION: str = _tv.TEMPLATE_VERSION

CAN_IDS: tuple[str, ...] = (
    "0x1A960004",
    "0x1A970004",
    "0x1A980004",
    "0x1A990004",
    "0x1A9A0004",
    "0x1A9B0004",
)
MAX_ROWS = 3200
SAMPLES = 500
FONT = Font(name="Calibri", size=11, color="404040")
HEADER_FONT = Font(name="Calibri", bold=True, size=10)
UI_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
UI_SUBTITLE_FONT = Font(name="Calibri", size=10, color="595959")
UI_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ACTION_PANEL_FILL = PatternFill(start_color="F6F8FA", end_color="F6F8FA", fill_type="solid")
ACTION_LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="24292F")
ACTION_HINT_FONT = Font(name="Calibri", size=9, color="57606A")
_THIN_BORDER_SIDE = Side(style="thin", color="D0D7DE")
ACTION_PANEL_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)


def _style_header(ws: Worksheet, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _int16_formula(b0_col: str, b1_col: str, row: int) -> str:
    raw = f"{b0_col}{row}*256+{b1_col}{row}"
    return f'=IF(Raw!A{row}="","",IF({raw}>=32768,{raw}-65536,{raw}))'


def _apply_action_panel_style(
    ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = ACTION_PANEL_FILL
            cell.border = ACTION_PANEL_BORDER


def _write_template_meta(ws: Worksheet) -> None:
    """Hidden column G + named ranges for VBA (TemplateVersion, etc.)."""
    ws[META_VERSION_CELL] = TEMPLATE_VERSION
    ws[META_RELEASE_DATE_CELL] = TEMPLATE_RELEASE_DATE
    ws[META_SUMMARY_CELL] = TEMPLATE_RELEASE_SUMMARY
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["G"].width = 2


def _define_template_names(wb: Workbook) -> None:
    from openpyxl.workbook.defined_name import DefinedName

    def ref(cell: str) -> str:
        letters = "".join(ch for ch in cell if ch.isalpha())
        digits = "".join(ch for ch in cell if ch.isdigit())
        return f"Instructions!${letters}${digits}"

    wb.defined_names.add(DefinedName("TemplateVersion", attr_text=ref(META_VERSION_CELL)))
    wb.defined_names.add(
        DefinedName("TemplateReleaseDate", attr_text=ref(META_RELEASE_DATE_CELL))
    )
    wb.defined_names.add(
        DefinedName("TemplateReleaseSummary", attr_text=ref(META_SUMMARY_CELL))
    )


def _write_instructions(ws: Worksheet) -> None:
    _write_template_meta(ws)

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "CAN Fault Recording"
    title.font = UI_TITLE_FONT
    title.fill = UI_HEADER_FILL
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws["A2"] = (
        f"v{TEMPLATE_VERSION} · released {TEMPLATE_RELEASE_DATE} · "
        f"{SUPPORTED_EXCEL_NOTE} · macros required"
    )
    ws["A2"].font = UI_SUBTITLE_FONT

    ws.merge_cells("D2:F2")
    ws["D2"] = "Import recording"
    ws["D2"].font = ACTION_LABEL_FONT
    ws["D2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("D5:F5")
    ws["D5"] = "Select a CAN fault recording .csv file"
    ws["D5"].font = ACTION_HINT_FONT
    ws["D5"].alignment = Alignment(horizontal="left", vertical="top", indent=1)
    _apply_action_panel_style(ws, 2, 5, 4, 6)

    lines: list[str] = [
        "",
        "About this template",
        f"Version {TEMPLATE_VERSION}",
        f"Released {TEMPLATE_RELEASE_DATE}",
        TEMPLATE_RELEASE_SUMMARY,
        *(f"· {note}" for note in TEMPLATE_RELEASE_NOTES),
        "",
        "Quick start",
        "1. Enable macros when opening this workbook.",
        "2. Use Import CSV (panel top-right) and choose a recording file.",
        "   Comment lines (#…) and the can_id header row are skipped automatically.",
        "3. Re-import replaces all previous Raw data.",
        "4. Watch the status bar at the bottom for import progress.",
        "5. If import fails, open sheet ImportLog for details.",
        "",
        "Sheets",
        "· Parsed — int16 channels filled by macro",
        "· Plot_* — four charts per CAN ID (500 points each)",
        "· Alt+F8 → RebuildAllPlotCharts if charts look wrong after import",
        "",
        "int16 = high byte × 256 + low byte (signed). Example: 0x01,0xF6 → 502.",
        "Grouping is by can_id in file order (not fixed 6-row blocks).",
    ]
    section_titles = {"About this template", "Quick start", "Sheets"}
    start_row = 3
    for offset, text in enumerate(lines):
        row = start_row + offset
        cell = ws.cell(row=row, column=1, value=text)
        if text in section_titles:
            cell.font = Font(name="Calibri", bold=True, size=11, color="24292F")
        elif text.startswith("Version ") or text.startswith("Released "):
            cell.font = Font(name="Calibri", size=10, color="595959")
        elif text == TEMPLATE_RELEASE_SUMMARY:
            cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
        elif text.startswith("· "):
            cell.font = Font(name="Calibri", size=10, color="404040")
        else:
            cell.font = FONT

    ws.column_dimensions["A"].width = 68
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 18


def _write_import_log(ws: Worksheet) -> None:
    headers = ("time", "step", "error_number", "error_message", "csv_path", "error_source")
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header(ws, 1, len(headers))
    ws["A2"] = "(empty until an import fails)"
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22


def _write_raw(ws: Worksheet) -> None:
    """Headers only; rows are filled exclusively by Import CSV macro."""
    headers = ["can_id", "b0", "b1", "b2", "b3", "b4", "b5", "b6", "b7"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header(ws, 1, len(headers))
    ws["A2"] = "(empty — use Import CSV...)"
    ws["A2"].font = Font(name="Arial", italic=True, color="808080")
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_parsed(ws: Worksheet) -> None:
    headers = [
        "row",
        "can_id",
        "seq_in_id",
        "b0",
        "b1",
        "b2",
        "b3",
        "b4",
        "b5",
        "b6",
        "b7",
        "b0b1",
        "b2b3",
        "b4b5",
        "b6b7",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header(ws, 1, len(headers))

    for row in range(2, MAX_ROWS + 2):
        ws.cell(row=row, column=1, value="=ROW()-1")
        ws.cell(row=row, column=2, value=f'=IF(Raw!A{row}="","",Raw!A{row})')
        ws.cell(row=row, column=3, value="")
        for byte_idx, raw_col in enumerate(range(2, 10), start=4):
            letter = get_column_letter(raw_col - 1)
            ws.cell(
                row=row,
                column=byte_idx,
                value=f'=IF(Raw!A{row}="","",HEX2DEC(MID(Raw!{letter}{row},3,2)))',
            )
        ws.cell(row=row, column=12, value=_int16_formula("D", "E", row))
        ws.cell(row=row, column=13, value=_int16_formula("F", "G", row))
        ws.cell(row=row, column=14, value=_int16_formula("H", "I", row))
        ws.cell(row=row, column=15, value=_int16_formula("J", "K", row))

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _write_plot_sheet(ws: Worksheet, can_id: str) -> None:
    ws["A1"] = "CAN ID"
    ws["B1"] = can_id
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT

    headers = ["sample", "b0b1", "b2b3", "b4b5", "b6b7"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=name)
    _style_header(ws, 2, len(headers))

    for sample in range(1, SAMPLES + 1):
        row = sample + 2
        ws.cell(row=row, column=1, value=sample)
        for ch_col in range(2, 6):
            ws.cell(row=row, column=ch_col, value="")

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 12
    # Charts are created by VBA (RebuildAllPlotCharts) — one line, 500 points each.


def _write_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Select CAN ID (open matching Plot_* sheet for charts)"
    ws["A1"].font = HEADER_FONT
    ws["B1"] = CAN_IDS[0]
    dv = DataValidation(type="list", formula1=f'"{",".join(CAN_IDS)}"', allow_blank=False)
    dv.add(ws["B1"])
    ws.add_data_validation(dv)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18

    ws["A3"] = "Plot sheets"
    ws["A3"].font = HEADER_FONT
    for idx, can_id in enumerate(CAN_IDS, start=4):
        short = can_id.replace("0x", "")
        ws.cell(row=idx, column=1, value=f"Plot_{short}")
        ws.cell(row=idx, column=2, value="Charts populate after Import CSV")


def _vba_bas_path() -> Path:
    return Path(__file__).parent / "can_fault_import.bas"


_WORKBOOK_OPEN_VBA = """\
Private Sub Workbook_Open()
    If Not EnsureExcelVersion() Then Exit Sub
    Call ProtectRawSheet
    Call PromptImportIfEmpty
End Sub
"""


def _excel_rgb(red: int, green: int, blue: int) -> int:
    """VBA/Excel RGB: R + G*256 + B*65536."""
    return int(red + green * 256 + blue * 65536)


class _ExcelRange(Protocol):
    @property
    def Left(self) -> float: ...

    @property
    def Top(self) -> float: ...

    @property
    def Width(self) -> float: ...

    @property
    def Height(self) -> float: ...

    def Select(self) -> None: ...


class _ExcelColor(Protocol):
    RGB: int


class _ExcelFill(Protocol):
    ForeColor: _ExcelColor


class _ExcelLine(Protocol):
    Visible: int
    ForeColor: _ExcelColor


class _ExcelShape(Protocol):
    Name: str

    def Delete(self) -> None: ...


class _ExcelShapes(Protocol):
    def __iter__(self) -> Iterator[_ExcelShape]: ...

    def AddShape(
        self, shape_type: int, left: float, top: float, width: float, height: float
    ) -> _ExcelButtonShape: ...


class _ExcelFont(Protocol):
    Name: str
    Bold: bool
    Size: float
    Fill: _ExcelFill


class _ExcelTextRange(Protocol):
    Text: str
    Font: _ExcelFont


class _ExcelTextFrame2(Protocol):
    TextRange: _ExcelTextRange
    VerticalAnchor: int
    HorizontalAnchor: int
    MarginLeft: float
    MarginRight: float
    MarginTop: float
    MarginBottom: float


class _ExcelButtonShape(Protocol):
    Name: str
    OnAction: str
    Fill: _ExcelFill
    Line: _ExcelLine
    TextFrame2: _ExcelTextFrame2


class _ExcelWindow(Protocol):
    DisplayGridlines: bool


class _ExcelApplication(Protocol):
    ActiveWindow: _ExcelWindow


class _ExcelWorksheetCom(Protocol):
    Shapes: _ExcelShapes

    def Range(self, cell1: str, cell2: str | None = None) -> _ExcelRange: ...

    def Activate(self) -> None: ...

    @property
    def Application(self) -> _ExcelApplication: ...


class _ExcelCustomProperty(Protocol):
    Value: str


class _ExcelCustomProperties(Protocol):
    def Item(self, name: str) -> _ExcelCustomProperty: ...

    def Add(self, name: str, link_to_content: bool, property_type: int, value: str) -> None: ...


class _ExcelWorkbookCom(Protocol):
    CustomDocumentProperties: _ExcelCustomProperties

    def Worksheets(self, name: str) -> _ExcelWorksheetCom: ...


def _add_import_csv_button(ws: _ExcelWorksheetCom) -> None:
    """Primary action button in the Instructions action panel."""
    for shape in list(ws.Shapes):
        if shape.Name == "btnImportCsv":
            shape.Delete()
    anchor = ws.Range("D3:F4")
    left = float(anchor.Left)
    top = float(anchor.Top)
    width = float(anchor.Width)
    height = float(anchor.Height)
    # msoShapeRoundedRectangle = 5
    btn = ws.Shapes.AddShape(5, left, top, width, height)
    btn.Name = "btnImportCsv"
    btn.OnAction = "ImportCanFaultCsv"
    btn.Fill.ForeColor.RGB = _excel_rgb(0, 120, 212)
    btn.Line.Visible = 0
    tf = btn.TextFrame2
    tf.TextRange.Text = "Import CSV…"
    tf.TextRange.Font.Name = "Calibri"
    tf.TextRange.Font.Bold = True
    tf.TextRange.Font.Size = 12
    tf.TextRange.Font.Fill.ForeColor.RGB = _excel_rgb(255, 255, 255)
    tf.VerticalAnchor = 3  # msoAnchorMiddle
    tf.HorizontalAnchor = 2  # msoAnchorCenter
    tf.MarginLeft = 6
    tf.MarginRight = 6
    tf.MarginTop = 4
    tf.MarginBottom = 4


def _upsert_custom_property(wb: _ExcelWorkbookCom, name: str, value: str) -> None:
    props = wb.CustomDocumentProperties
    try:
        props.Item(name).Value = value
        return
    except Exception:
        pass
    # msoPropertyTypeString = 4
    props.Add(name, False, 4, value)


def _set_workbook_release_metadata(wb: _ExcelWorkbookCom) -> None:
    _upsert_custom_property(wb, "TemplateVersion", TEMPLATE_VERSION)
    _upsert_custom_property(wb, "TemplateReleaseDate", TEMPLATE_RELEASE_DATE)
    _upsert_custom_property(wb, "TemplateReleaseSummary", TEMPLATE_RELEASE_SUMMARY)


def _polish_instructions_sheet(ws: _ExcelWorksheetCom) -> None:
    """Hide gridlines and leave focus on content, not cell grid."""
    ws.Activate()
    ws.Application.ActiveWindow.DisplayGridlines = False
    ws.Range("A3").Select()


def _embed_vba_and_button(xlsx_path: Path, xlsm_path: Path) -> None:
    try:
        import win32com.client  # type: ignore[import-untyped]
    except ImportError as exc:
        msg = "pywin32 required for .xlsm build: pip install pywin32"
        raise RuntimeError(msg) from exc

    bas_path = _vba_bas_path()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        try:
            wb.CheckCompatibility = True
        except Exception:
            pass
        for component in list(wb.VBProject.VBComponents):
            if component.Name == "CanFaultImport":
                wb.VBProject.VBComponents.Remove(component)
        module = wb.VBProject.VBComponents.Import(str(bas_path.resolve()))
        module.Name = "CanFaultImport"

        for component in list(wb.VBProject.VBComponents):
            if component.Type == 100 and component.Name == "ThisWorkbook":
                line_count = component.CodeModule.CountOfLines
                if line_count > 0:
                    component.CodeModule.DeleteLines(1, line_count)
                component.CodeModule.AddFromString(_WORKBOOK_OPEN_VBA)
                break

        instructions = cast(_ExcelWorksheetCom, wb.Worksheets("Instructions"))
        _add_import_csv_button(instructions)
        _polish_instructions_sheet(instructions)
        _set_workbook_release_metadata(cast(_ExcelWorkbookCom, wb))

        wb.SaveAs(str(xlsm_path.resolve()), FileFormat=XLSM_FILE_FORMAT)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()


def build_template(out_path: Path) -> None:
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    ws_inst = wb.create_sheet("Instructions", 0)
    _write_instructions(ws_inst)

    ws_raw = wb.create_sheet("Raw", 1)
    _write_raw(ws_raw)

    ws_log = wb.create_sheet("ImportLog", 2)
    _write_import_log(ws_log)

    ws_parsed = wb.create_sheet("Parsed", 3)
    _write_parsed(ws_parsed)

    _write_dashboard(wb)

    for can_id in CAN_IDS:
        short = can_id.replace("0x", "")
        ws_plot = wb.create_sheet(f"Plot_{short}")
        _write_plot_sheet(ws_plot, can_id)

    _define_template_names(wb)

    notes_text = "; ".join(TEMPLATE_RELEASE_NOTES)
    wb.properties.title = "CAN Fault Recording Template"
    wb.properties.subject = TEMPLATE_RELEASE_SUMMARY
    wb.properties.description = (
        f"Version {TEMPLATE_VERSION}, released {TEMPLATE_RELEASE_DATE}. {notes_text}"
    )
    wb.properties.keywords = f"CAN;fault;recording;v{TEMPLATE_VERSION}"
    wb.properties.category = "McuCanMap"
    wb.properties.creator = "McuCanMap"
    wb.save(out_path)


def main() -> None:
    base = Path(__file__).parent
    xlsx_path = base / "can_fault_recording_template.xlsx"
    xlsm_path = base / "can_fault_recording_template.xlsm"

    build_template(xlsx_path)
    print(f"Wrote: {xlsx_path}  (headers only, no sample data)")
    print(f"  Template v{TEMPLATE_VERSION} · released {TEMPLATE_RELEASE_DATE}")

    try:
        _embed_vba_and_button(xlsx_path, xlsm_path)
        print(f"Wrote: {xlsm_path}  ({SUPPORTED_EXCEL_NOTE}, use this file)")
        print(f"  Template v{TEMPLATE_VERSION} · released {TEMPLATE_RELEASE_DATE}")
    except Exception as exc:
        print(f"WARNING: could not build .xlsm ({exc})")
        print("  Close Excel, then run setup_and_build_excel_template.ps1")


if __name__ == "__main__":
    main()
