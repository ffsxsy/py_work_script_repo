"""Smoke tests for CAN fault Excel template build (openpyxl only, no Excel COM)."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from types import ModuleType

from openpyxl import load_workbook

_FAULT_DIR = Path(__file__).resolve().parents[1] / "1.fault_recording_parse_excel_template"


def _load_fault_module(stem: str) -> ModuleType:
    path = _FAULT_DIR / f"{stem}.py"
    module_name = f"_fault_test_{stem}"
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        msg = f"cannot load module from {path}"
        raise ImportError(msg)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


_build = _load_fault_module("build_can_fault_excel_template")
_instructions = _load_fault_module("instructions_content")
build_template = _build.build_template
INSTRUCTIONS_LAYOUT = _instructions.INSTRUCTIONS_LAYOUT


def test_build_template_instructions_layout(tmp_path: Path) -> None:
    out = tmp_path / "can_fault_recording_template.xlsx"
    build_template(out)

    wb = load_workbook(out)
    try:
        assert "Instructions" in wb.sheetnames
        ws = wb["Instructions"]
        layout = INSTRUCTIONS_LAYOUT

        assert ws.freeze_panes == layout.freeze_panes
        assert ws["A1"].value == layout.title
        assert ws["A3"].value == "About this template"
        assert ws.row_dimensions[layout.gap_row].height == layout.section_gap_height
        tab_color = ws.sheet_properties.tabColor
        assert tab_color is not None
        assert tab_color.rgb is not None
        assert tab_color.rgb.endswith(layout.tab_color)

        for name in ("Raw", "ImportLog", "Parsed", "Dashboard"):
            assert name in wb.sheetnames
        assert any(name.startswith("Plot_") for name in wb.sheetnames)
    finally:
        wb.close()


def test_build_template_version_metadata(tmp_path: Path) -> None:
    out = tmp_path / "meta.xlsx"
    build_template(out)

    wb = load_workbook(out)
    try:
        props = wb.properties
        assert props.title == "CAN Fault Recording Template"
        assert props.description is not None
        assert "1.0.1" in props.description
    finally:
        wb.close()
